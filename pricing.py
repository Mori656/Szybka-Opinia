import datetime
from openpyxl.styles import numbers, Font, Alignment, PatternFill, Border, Side
import os
import shutil

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.workbook.properties import CalcProperties

class Pricing:
    def __init__(self, info, device, pricing_file_name):
        self.device = device
        self.info = info

        folder_name = f"{'Wygenerowane_opinie/' + self.info['Nazwa pliku']}/Wycena_{self.device.info['Numer ewidencyjny']}"
        os.mkdir(folder_name)
        path_to_folder = folder_name + '/'
        self.pricing_file_name = f"{path_to_folder}{pricing_file_name}"
        self.pricing_page_path = "Cennik.pdf"      
        shutil.copy2(self.pricing_page_path, path_to_folder)

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        ft1 = Font(name = "Trebuchet MS", size=11)
        ft_title = Font(name = "Trebuchet MS", size=18)
        fill_color = PatternFill(start_color="ffff00", end_color="ffff00", fill_type="solid")
        ft_link = Font(color="0563C1", underline="single")

        wb = Workbook()
        ws_form = wb.active
        ws_form.title = "Formularz"
        ws_form.column_dimensions['A'].width = 8.43
        ws_form.column_dimensions['B'].width = 23.86
        ws_form.column_dimensions['C'].width = 10.43
        ws_form.column_dimensions['D'].width = 8.43
        ws_form.column_dimensions['E'].width = 8.43
        ws_form.column_dimensions['F'].width = 5.43
        ws_form.column_dimensions['G'].width = 18.14
        ws_form.row_dimensions[31].height = 49.5
        ws_form.row_dimensions[33].height = 51.75
        ws_form.row_dimensions[35].height = 53.25

        ws_pricing = wb.create_sheet("Wycena oraz linki")
        ws_pricing.column_dimensions['A'].width = 23.86
        ws_pricing.column_dimensions['B'].width = 23.86
        ws_pricing.column_dimensions['C'].width = 19.43
        ws_pricing.column_dimensions['D'].width = 49.0
        ws_pricing.row_dimensions[2].height = 33.0
        ws_pricing.row_dimensions[18].height = 30.0
        ws_pricing.row_dimensions[19].height = 30.0
        ws_pricing.row_dimensions[29].height = 60.0

        def merge_and_write(ws, text, start_cell, end_cell=None, price=False, format=False, link=False):
            if end_cell:
                ws.merge_cells(f"{start_cell}:{end_cell}")

            # Wstaw wartość
            if isinstance(text, str) and text.startswith("="):  # formuła
                ws[start_cell] = text
            elif price and text != "":
                try:
                    ws[start_cell] = float(text)
                except ValueError:
                    ws[start_cell] = text  # fallback
            else:
                ws[start_cell] = text

            if link:
                ws[start_cell].hyperlink = text
                ws[start_cell].style = "Hyperlink"
                ws[start_cell].font = ft_link

            # Ustaw format walutowy
            if price or format:
                ws[start_cell].number_format = '#,##0.00" zł"'

        def exploatation_time_protocol(time):
            if time == 1:
                return "1 rok"
            elif time % 10 in [2,3,4]:
                return f"{time} lata"
            else:
                return f"{time} lat"

        merge_and_write(ws_form, "Departament Korporacyjny Informatyki", "A2", "D2")
        merge_and_write(ws_form, f"Tarnów, dn. {datetime.date.today().strftime('%d.%m.%Y')}", "E2", "G2")
        merge_and_write(ws_form, "Zespół Serwisu IT", "A3", "B3")
        merge_and_write(ws_form, "Wycena sprzętu komputerowego", "A6", "G7")
        merge_and_write(ws_form, f"nr {self.device.info['Numer wyceny']}/{datetime.date.today().strftime('%Y')} "
                            f"do opinii nr {self.info['Numer opinii']}/{datetime.date.today().strftime('%Y')}", "A9", "G9")

        merge_and_write(ws_form, "Nazwa sprzętu z MPK użytkownika", "A13", "B13")
        merge_and_write(ws_form, f"{self.device.info['MPK']}", "C13", "G13")

        merge_and_write(ws_form, "NR SAP/SAT", "A14", "B14")
        merge_and_write(ws_form, f"{self.device.info['Numer środka trwałego']}", "C14", "G14")

        merge_and_write(ws_form, "NR Inw.", "A15", "B15")
        merge_and_write(ws_form, f"{self.device.info['Nr inwentarzowy']}", "C15", "G15")

        merge_and_write(ws_form, "Marka:", "A17", "B17")
        merge_and_write(ws_form, f"{self.device.info['Producent']}", "C17", "G17")

        merge_and_write(ws_form, "Model:", "A18", "B18")
        merge_and_write(ws_form, f"{self.device.info['Model']}", "C18", "G18")

        merge_and_write(ws_form, "Typ:", "A19", "B19")
        merge_and_write(ws_form, f"{self.device.info['Typ']}", "C19", "G19")

        merge_and_write(ws_form, "Nr Seryjny:", "A20", "B20")
        merge_and_write(ws_form, f"{self.device.info['Numer seryjny']}", "C20", "G20")

        merge_and_write(ws_form, "Okres eksploatacji", "A21", "B21")
        exploatation_time = exploatation_time_protocol(int(datetime.date.today().strftime('%Y')) - int(datetime.datetime.strptime(device.info['Data zakupu'], '%d.%m.%Y').year))
        merge_and_write(ws_form, f"{exploatation_time}", "C21", "G21")

        merge_and_write(ws_form, "Parametry:", "A22", "B22")
        merge_and_write(ws_form, "Procesor:", "B23")
        merge_and_write(ws_form, f"{self.device.info['Model procesora']}", "C23", "G23")
        merge_and_write(ws_form, "Płyta główna:", "B24")
        merge_and_write(ws_form, f"{self.device.info['Producent']}", "C24", "G24")
        merge_and_write(ws_form, "RAM:", "B25")
        merge_and_write(ws_form, f"{self.device.info['Pamięć RAM (GB)']} GB", "C25", "G25")
        merge_and_write(ws_form, "Grafika:", "B26")
        merge_and_write(ws_form, f"{self.device.info['Grafika']}", "C26", "G26")
        merge_and_write(ws_form, "Dysk:", "B27")
        merge_and_write(ws_form, f"{self.device.info['Pamięć HDD (GB)']} GB", "C27", "G27")
        merge_and_write(ws_form, "System operacyjny:", "B28")
        merge_and_write(ws_form, f"{self.device.info['System operacyjny']}", "C28", "G28")

        merge_and_write(ws_form, "Stan techniczny:", "A30", "B30")
        merge_and_write(ws_form, f"{self.device.info['Stan']}", "C30", "G30")

        merge_and_write(ws_form, "Uwagi do stanu technicznego:", "A31", "B31")
        merge_and_write(ws_form, f"{self.device.info['Wymiana']}", "C31", "G31")

        merge_and_write(ws_form, "Wycena rynkowa netto sprzętu \n(o takiej samej lub podobnej konfiguracji):", "A33", "B33")
        #='Wycena oraz linki'!B14
        merge_and_write(ws_form, "='Wycena oraz linki'!B14", "C33", "G33", format=True)

        merge_and_write(ws_form, "Wycena rynkowa netto po uwzględnieniu uwag dot. stanu technicznego", "A35", "B35")
        #='Wycena oraz linki'!B29
        merge_and_write(ws_form, "='Wycena oraz linki'!B29", "C35", "G35", format=True)

        merge_and_write(ws_form, "Wycenił:", "A38", "B38")
        merge_and_write(ws_form, "Sprawdził:", "C38", "D38")
        merge_and_write(ws_form, "Zatwierdził:", "F38", "G38")

        merge_and_write(ws_pricing, "Wycena rynkowa netto sprzętu (od takiej samej lub podobnej konfiguracji):", "A2", "C2")
        merge_and_write(ws_pricing, "Linki do wyceny sprzętu", "D2")

        for i in range(1, 11):
            merge_and_write(ws_pricing, i, f"A{i+2}")
            merge_and_write(ws_pricing, f"{self.device.info[f'Price{i}']}", f"B{i+2}", price=True, format=True)
            merge_and_write(ws_pricing, f"{self.device.info[f'Link{i}']}", f"D{i+2}", link=True)
        
        merge_and_write(ws_pricing, "Średnia brutto", "A13")
        merge_and_write(ws_pricing, '=IFERROR(AVERAGEIF(B3:B12,"<>"), 0)', "B13", format=True)
        merge_and_write(ws_pricing, "Średnia netto", "A14",)
        ws_pricing["A14"].fill = fill_color

        merge_and_write(ws_pricing, '=IFERROR(B13/1.23, 0)', "B14", format=True)
        ws_pricing["B14"].fill = fill_color

        merge_and_write(ws_pricing, "Wycena uszkodzonych elementów", "A18", "B18")
        merge_and_write(ws_pricing, "Nazwa elementu/usługi", "C18")
        merge_and_write(ws_pricing, "Linki do wyceny uszkodzonych elementów", "D18")

        i = 1
        if device.info["Cena_cooling"] != "":
            merge_and_write(ws_pricing, i, f"A{18+i}")
            merge_and_write(ws_pricing, f"{device.info['Cena_cooling']}", f"B{18+i}", price=True, format=True)
            merge_and_write(ws_pricing, "Czyszczenie układu chłodzenia", f"C{18+i}")
            merge_and_write(ws_pricing, f"{device.info['Link_cooling']}", f"D{18+i}", link=True)
            i = i + 1

        if device.info["Cena_battery"] != "":
            merge_and_write(ws_pricing, i, f"A{18+i}")
            merge_and_write(ws_pricing, f"{device.info['Cena_battery']}", f"B{18+i}", price=True, format=True)
            merge_and_write(ws_pricing, "Koszt baterii", f"C{18+i}")
            merge_and_write(ws_pricing, f"{device.info['Link_battery']}", f"D{18+i}", link=True)

        merge_and_write(ws_pricing, "Suma brutto", "A25")
        merge_and_write(ws_pricing, "Suma netto", "A26")
        ws_pricing["A26"].fill = fill_color

        merge_and_write(ws_pricing, '=IFERROR(SUM(B19:B24), 0)', "B25", format=True)
        merge_and_write(ws_pricing, '=IFERROR(B25/1.23, 0)', "B26", format=True)
        ws_pricing["B26"].fill = fill_color

        merge_and_write(ws_pricing, "Wycena rynkowa netto po uwzględnieniu uwag dot. stanu technicznego i różnic w konfiguracji", "A29")
        merge_and_write(ws_pricing, "=B14-B26", "B29", format=True)
        ws_pricing["A29"].fill = fill_color
        ws_pricing["B29"].fill = fill_color

        for col in ws_form.columns:
            for cell in col:
                cell.font = ft1
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(wrap_text=True)

        for col in ws_pricing.columns:
            for cell in col:
                if isinstance(cell.value, str):
                    cell.alignment = Alignment(wrap_text=True)

        for col in [3,5,6]:  # kolumny A-F
            col_letter = get_column_letter(col)
            for row in range(1,50):
                cell = ws_form[f"{col_letter}{row}"]
                cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in [4]:  # kolumny A-F
            col_letter = get_column_letter(col)
            for row in range(1,50):
                cell = ws_pricing[f"{col_letter}{row}"]
                cell.alignment = Alignment(horizontal='left')

        for row in ws_pricing["A2:D14"]:
            for cell in row:
                cell.border = thin_border

        for row in ws_pricing["A18:D26"]:
            for cell in row:
                cell.border = thin_border

        for row in ws_pricing["A29:B29"]:
            for cell in row:
                cell.border = thin_border

        ws_form["A6"].font = ft_title
        ws_form["A6"].alignment = Alignment(horizontal='center', vertical='center')
        ws_form["A9"].alignment = Alignment(horizontal='center', vertical='center')
        ws_form["A13"].alignment = Alignment(horizontal='left')
        ws_form["C31"].alignment = Alignment(horizontal='center', wrap_text=True)

        wb.save(f"{self.pricing_file_name}.xlsx")