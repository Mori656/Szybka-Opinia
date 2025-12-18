import datetime
import re
import tkinter as tk
from tkinter import ttk, messagebox

import openpyxl

from Classes.betterComboBox import BetterComboBox
from Classes.device import Device
from Classes.request import Request

from style import apply_custom_style


class Form:
    list_devices = []
    info = {
        "Numer opinii": None,
        "Zglaszajacy": None,
        "Dla kogo": None,
        "Autor": None,
        "Nazwa pliku": None,
    }
    departments_path = "./Bazy_danych/Departments.xlsx"
    departments_codes_column = "Kod działu"
    departments_names_column = "Nazwa działu"

    error_count = 0

    def __init__(self):
        self.window = tk.Tk()
        self.window.title("Karol Mielony")
        self.window.geometry("1280x720")

        self.window.state("normal")
        self.window.configure(bg="#1e1e2f")

        self.window.grid_rowconfigure(0, weight=1)
        self.window.grid_columnconfigure(0, weight=1)

        ### Styl
        apply_custom_style()

        # Główna ramka podzielona na lewo/prawo
        self.main_frame = tk.Frame(self.window, bg="#1e1e2f")
        self.main_frame.grid(row=0,column=0,sticky='nsew')
        # self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)

        # Lewa ramka – urządzenia (ze scrollowaniem)
        self.left_frame = tk.Frame(self.main_frame, bg="#1e1e2f")
        self.left_frame.pack(side="left", fill="both", expand=True, padx=(0, 20))

        self.label_devices = ttk.Label(self.left_frame, text="Urządzenia:")
        self.label_devices.pack(anchor="w", padx=5, pady=(0, 5))

        # Stwórz canvas i scrollbar
        self.devices_canvas = tk.Canvas(self.left_frame, bg="#1e1e2f", highlightthickness=0)
        self.devices_canvas.pack(side="left", fill="both", expand=True)

        self.scrollbar = ttk.Scrollbar(self.left_frame, orient="vertical", command=self.devices_canvas.yview)
        self.scrollbar.pack(side="right", fill="y")

        self.devices_canvas.configure(yscrollcommand=self.scrollbar.set)

        # Kontener wewnątrz canvas
        self.devices_container = tk.Frame(self.devices_canvas, bg="#1e1e2f")
        self.container_window = self.devices_canvas.create_window((0, 0), window=self.devices_container, anchor="nw")

        # Aktualizacja scrollregion
        self.devices_container.bind(
            "<Configure>",
            lambda e: self.devices_canvas.configure(scrollregion=self.devices_canvas.bbox("all"))
        )

        # Dynamiczne dopasowanie szerokości kontenera
        def resize_canvas(event):
            self.devices_canvas.itemconfig(self.container_window, width=event.width)

        self.devices_canvas.bind("<Configure>", resize_canvas)

        # Obsługa scrolla myszką
        def _on_mousewheel(event):
            self.devices_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self.devices_canvas.bind_all("<MouseWheel>", _on_mousewheel)
        self.devices_canvas.bind_all("<Button-4>", lambda e: self.devices_canvas.yview_scroll(-1, "units"))
        self.devices_canvas.bind_all("<Button-5>", lambda e: self.devices_canvas.yview_scroll(1, "units"))

        self.add_device()

        # Prawa ramka – formularz opinii
        self.right_frame = tk.Frame(self.main_frame, bg="#1e1e2f")
        self.right_frame.pack(side="left", fill="both", expand=True)

        self.create_input("Numer opinii: ", "opinion_number", self.right_frame, only_int=True)
        self.create_input("Zgłaszający: ", "reporter", self.right_frame)

        label = ttk.Label(self.right_frame, text="Dział:")
        label.pack(pady=(15, 5), anchor="w", padx=10)
        options = self.get_departments()
        combo = BetterComboBox(self.right_frame,
                               completevalues=options,
                               listbox_bg="#2e2e3e",
                               listbox_fg="white")
        combo.pack(ipady=3, ipadx=3, fill="x", padx=10)
        combo.bind("<Return>", lambda e: combo.event_generate("<Down>"))
        combo.bind("<Escape>", lambda e: combo.selection_clear())
        self.bind_mousewheel(self.left_frame, self.devices_canvas)
        setattr(self, f"field_for_who", combo)

        self.create_input("Autor: ", "author", self.right_frame)

        # Ramka tylko na przyciski (w poziomie)
        self.button_frame = tk.Frame(self.right_frame, bg="#1e1e2f")
        self.button_frame.pack(pady=20)  # to ustala wysokość tej sekcji w pionowym układzie

        # Przyciski wewnątrz tej ramki – poziomo
        self.button_add_devices = ttk.Button(self.button_frame, text="Dodaj urządzenie", command=self.add_device)
        self.button_add_devices.pack(side="left", padx=10)
        
        self.button_generate = ttk.Button(self.button_frame, text="Generuj", command=self.generatePage)
        self.button_generate.pack(side="left", padx=10)
        

        self.button_next = ttk.Button(self.button_frame, text="Dalej", command=self.nextPage)
        self.button_next.pack(side="left", padx=10)
        self.button_next.config(state="disabled")

        # Page 2 ( Request )
        self.request = Request(self.window, self.main_frame)
        self.main_frame.tkraise()

    def on_close(self):
        self.window.unbind_all("<MouseWheel>")
        self.window.unbind_all("<Button-4>")
        self.window.unbind_all("<Button-5>")

        self.window.destroy()

    def create_input(self, label_text, attr_name, parent, only_int=False):
        label = ttk.Label(parent, text=label_text)
        label.pack(pady=(15, 5), anchor="w", padx=10)

        entry = ttk.Entry(parent)

        if only_int:
            def validate_int(P):
                if P == "":
                    return True
                return P.isdigit()

            vcmd = parent.register(validate_int)
            entry.config(validate="key", validatecommand=(vcmd, "%P"))

        entry.pack(ipady=3, ipadx=3, fill="x", padx=10)
        setattr(self, f"field_{attr_name}", entry)

    def add_device(self):
        new_device = Device(self.devices_container, len(self.list_devices) + 1, parent=self)
        self.list_devices.append(new_device)

    def renumber_devices(self):
        for idx, device in enumerate(self.list_devices, start=1):
            device.label_lp.config(text=f"Urządzenie {idx}:")


    def get_departments(self):
        wb = openpyxl.load_workbook(self.departments_path)
        sheet = wb.active
        options = []
        d_full_name = ""
        for index, cell in enumerate(sheet[1]):
            col_index = index + 1
            key = str(cell.value)
            #Jesli nazwa kolumy jest zgodna z szukaną
            if key == self.departments_codes_column:
                #Przejdz przez instniejące wiersze
                for i in range(2, sheet.max_row + 1):
                    #Spradz czy jest wartość w komórce
                    if sheet.cell(row=i, column=col_index).value:
                        #Jesli jest warość to przejdz po wierszu wstecz i znajdz nazwe dzialu
                        for y in range(i):
                            if sheet.cell(row=i-y, column=1).value:
                                d_code_master = str(sheet.cell(row=i-y, column=1).value)
                                d_name_master = str(sheet.cell(row=i-y, column=2).value)
                                break
                        master_part = f"[{d_code_master}] {d_name_master}"

                        if col_index != 1:
                            d_code = str(sheet.cell(row=i, column=col_index).value)
                            d_name = str(sheet.cell(row=i, column=col_index + 1).value)
                            d_full_name = f"{master_part} / [{d_code}] {d_name}"
                        else:
                            d_full_name = master_part
                        options.append(d_full_name)
        return options

    def format_reporter_name(self, full_name):
        parts = full_name.strip().split()

        if len(parts) >= 2:
            initial = parts[0][0].upper()
            surname = parts[-1].capitalize()
            return f"{initial}.{surname}"
        else:
            return full_name

    def format_author_name(self, full_name):
        parts = full_name.strip().split()

        if len(parts) >= 2:
            initial = parts[0][0].upper()
            surname = parts[1][0].upper()
            return f"{initial}{surname}"
        else:
            return full_name

    def data_validation(self):
        # Sprawdź, czy każde urządzenie ma wypełnione ID
        for idx, device in enumerate(self.list_devices):
            if not device.field_id.get().strip():
                self.error_count += 1
                if self.error_count == 5:
                    messagebox.showerror(f"Znowu ten bład?", f"Proszę sprawdź najpierw czy wspiałeś odpowiedni numer  \nಠ‿ಠ ")
                elif self.error_count == 10:
                    messagebox.showerror(f"Znowu to samo?", f"Serio wystarczy że wpiszesz jakąkolwiek cyfrę i będzie dobrze. \n(ಠ_ಠ) ")
                elif self.error_count == 15:
                    messagebox.showerror(f"Jesteś niepoważny?", f"Skończyło się bycie miłym to jest twoja ostatnia szansa! \nಠ▃ಠ")
                elif self.error_count == 16:
                    messagebox.showerror(f"Koniec zabawy!", "Jeśli nie umiesz używać programu to nie będziesz go używał!  \n(ノಠ益ಠ)ノ")
                    self.window.destroy()
                    return False
                else:
                    messagebox.showerror(f"{self.field_author.get()} co ześ narobił?", f"Urządzenie {idx + 1} nie ma wpisanego odpowiednigo indetyfikatora.")
                return False

        # Lista wymaganych pól w formularzu
        missing_fields = []
        if not self.field_opinion_number.get().strip():
            missing_fields.append("Numer opinii")
        if not self.field_reporter.get().strip():
            missing_fields.append("Zgłaszający")
        if not self.field_for_who.get().strip():
            missing_fields.append("Dział")
        if not self.field_author.get().strip():
            missing_fields.append("Autor")

        # Jeśli jakieś pola są puste – zapytaj użytkownika
        if missing_fields:
            proceed = messagebox.askyesno(
                "Chyba o czymś zapomniałeś",
                f"Nie uzupełniono następujących pól:\n- " +
                "\n- ".join(missing_fields) +
                "\nBrak tych danych może skutkować błędnym działaniem programu." +
                "\n\nCzy chcesz kontynuować na własną odpowidzialność?"
            )
            return proceed

        return True

    def generatePage(self):
        if not self.data_validation():
            return  # zatrzymaj, jeśli walidacja się nie powiedzie

        for device in self.list_devices:
            device.findData()

        # Aktywowanie przycisku "Dalej"
        self.button_next.config(state="normal")

        # Pobranie świerzych informacji
        self.getInfo()
        self.request.document_info = self.info
        self.request.showRequests(list(self.list_devices))
        self.request.frame_request.tkraise()

    def nextPage(self):
        self.request.frame_request.tkraise()

    def departments_trim(self):
        department_name = self.field_for_who.get()
        # Pozbycie się skrótów
        trimmed = re.sub(r'\[[^\]]*\]\s*', '', department_name).replace('  ', ' ').strip()
        # Każde słowo dłuższe niż 1 litera: pierwsza litera duża, reszta mała
        def capitalize_word(word):
            return word[0].upper() + word[1:].lower() if len(word) > 1 else word
        trimmed = ' '.join(capitalize_word(w) for w in trimmed.split())
        return trimmed

    def getInfo(self):
        self.info["Numer opinii"] = self.field_opinion_number.get()
        self.info["Zglaszajacy"] = self.field_reporter.get()
        self.info["Dla kogo"] = self.departments_trim()
        self.info["Autor"] = self.field_author.get()

        reporter_name = self.format_reporter_name(self.info["Zglaszajacy"])
        author_name = self.format_author_name(self.info["Autor"])
        
        self.folder_name = f"{self.info['Numer opinii']}_{datetime.date.today().strftime('%Y')}_{author_name}_{reporter_name}"
        
        self.info["Nazwa pliku"] = self.folder_name

    def bind_mousewheel(self, widget, target_canvas):
        widget.bind("<Enter>", lambda e: target_canvas.bind_all("<MouseWheel>", lambda ev: target_canvas.yview_scroll(
            int(-1 * (ev.delta / 120)), "units")))
        widget.bind("<Leave>", lambda e: target_canvas.unbind_all("<MouseWheel>"))

