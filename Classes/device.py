import datetime
import tkinter as tk
import os
from dataclasses import fields
from tkinter import ttk, messagebox
from tkinter.ttk import Radiobutton, Checkbutton

import openpyxl

from Classes.betterComboBox import BetterComboBox

class Device:
    paths_folder = "./Bazy_danych/"
    paths = ['Ewidencja.xlsx', 'Log1.xlsx', 'Log2.xlsx', 'Log3.xlsx'] # Path to Log Export
    destroy_formula = ("Sprzęt zostaje zakwalifikowany do likwidacji ze "
                       "względu na niewspierany przez Microsoft system operacyjny, "
                       "podzespoły, które nie pozwalając na komfortową"
                       " pracę użytkownika z obecnie wykorzystywanymi "
                       "aplikacjami w Grupie Azoty S.A.. Dysk twardy "
                       "został wymontowany oraz wyczyszczono "
                       "jego zawartość. Proszę rozpocząć procedurę likwidacji sprzętu.")
    valuation_destroy_formula = ("Zalecane podstawowe czyszczenie układu chłodzenia i wymiana baterii. "
                                 "Przeinstalowano system operacyjny, dane z dysku usunięto. w przypadku gdy"
                                 "użytkownik nie zdecyduje się na odkup laptopa, proszę o przekazanie laptopa"
                                 "na majątek serwisu IT.")
    valuation_takeover_formula = ("Zalecane podstawowe czyszczenie układu chłodzenia i wymiana baterii. "
                                 "Przeinstalowano system operacyjny, dane z dysku usunięto. w przypadku gdy"
                                 "użytkownik nie zdecyduje się na odkup laptopa, proszę o rozpoczęcie procedury "
                                  "likwidacji.")
    state_good_formula = "Dobry, normalne ślady użytkowania"
    state_passable_formula = "Dostateczny, widoczne ślady użytkowania"
    change_cooling_and_battery_formula = "Należy wykonać czyszczenie układu chłodzenia oraz wymienić baterię"
    change_cooling_formula = "Należy wykonać czyszczenie układu chłodzenia"
    change_battery_formula = "Należy wymienić baterię"

    required_fields = {
            "Typ", "Producent", "Model", "Numer seryjny", "Data zakupu", "Numer ewidencyjny",
            "Numer środka trwałego", "System operacyjny", "Model procesora",
            "Pamięć RAM (GB)", "Pamięć HDD (GB)","Nr inwentarzowy"
        }
    search_keys = {"ID":"Numer ewidencyjny","SAP":"Numer środka trwałego","INV":"Nr inwentarzowy"}

    def __init__(self, frame, lp, parent=None):
        self.parent = parent

        self.outer_container = tk.Frame(frame, bg="#1e1e2f")
        self.outer_container.pack(padx=5, pady=(5, 0), fill="x", expand=True)


        self.label_lp = ttk.Label(self.outer_container, text=f"Urządzenie {lp}:", font=("Segoe UI", 10, "bold"))
        self.label_lp.pack(pady=(0, 5), anchor="w", padx=0)

        # "Card" dla jednego urządzenia
        self.container = tk.Frame(self.outer_container, bg="#2e2e3e", bd=0, highlightbackground="#666666", highlightcolor="white", highlightthickness=2)
        self.container.pack(padx=10, pady=5, fill="x", expand=True)

        self.container.grid_columnconfigure(0, weight=0)  # dla labela
        self.container.grid_columnconfigure(1, weight=1)

        # Etykieta po lewej
        self.label_id = tk.Label(
            self.container,
            text="Wprowadź",
            bg="#2e2e3e",
            fg="#ffffff"
        )
        self.label_id.grid(row=0, column=0, sticky="w", padx=10, pady=(10, 5))
        # Opcje wyszukiwania sprzętu
        options = ["ID","SAP","INV"]
        self.combo = BetterComboBox(self.container,
                               completevalues=options,
                               listbox_bg="#2e2e3e",
                               listbox_fg="white")
        self.combo.set(options[0])
        self.combo.grid(row=0, column=1, sticky="w", padx=10, pady=(10, 5))
        self.combo.bind("<Return>", lambda e: self.combo.event_generate("<Down>"))
        self.combo.bind("<Escape>", lambda e: self.combo.selection_clear())
        # Przycisk "usuń" po prawej

        destroy_button = ttk.Button(self.container,
                                    text="X",
                                    style="Destroy.TButton",
                                    command=self.destroy_self)
        destroy_button.grid(row=0, column=2, sticky="e", padx=10, pady=(10, 5))
        destroy_button.configure(width=2)

        # Pole entry na całą szerokość pod spodem
        self.field_id = ttk.Entry(self.container)
        self.field_id.grid(row=1, column=0, columnspan=3, sticky="ew", padx=10, pady=(0, 10), ipady=4)

        self.lp = lp
        self.info = {}
        self.pricing_file_name=None


    def destroy_self(self):
        self.outer_container.destroy()
        if self in self.parent.list_devices:
            self.parent.list_devices.remove(self)
            self.parent.renumber_devices()

    def findData(self):
        selected_key = self.combo.get()
        
        self.main_key = self.search_keys[selected_key]

        self.info = {key: "" for key in self.required_fields}
        self.info['Price'] = ""
        self.info['Link'] = ""
        self.info[self.main_key] = self.field_id.get()
        missing_fields = set(self.required_fields)
        missing_fields.discard(self.main_key)

        # Sprawdź, które wymagane pola nie zostały znalezione
        found_fields = self.findDataByKey(self.main_key)
        missing_fields -= found_fields

        for key in self.search_keys:
            search_field = self.search_keys[key]
            if search_field == self.main_key:
                continue
            if search_field in missing_fields:
                # Jeśli pole jest w missing_fields, pomijamy je
                continue
            found_fields = self.findDataByKey(search_field)
            missing_fields -= found_fields
        data_str = self.info.get("Data zakupu", "")
        try:
            parsed = datetime.datetime.strptime(data_str, "%Y-%m-%d %H:%M:%S")
            # Przechowuj w zmienionym formacie, jeśli chcesz
            self.info["Data zakupu"] = parsed.strftime("%d.%m.%Y")
        except (ValueError, TypeError):
            self.info["Data zakupu"] = "01.01.2000"
        
        self.info['Cena_cooling'] = 180
        self.info['Link_cooling'] = 'https://aidcom.eu/cennik/'
        

        if missing_fields:
            messagebox.showwarning(
                f"Brakujące dane urządzenia {self.lp}",
                "Nie znaleziono danych dla pól:\n\n" + "\n".join(sorted(missing_fields))
            )

    def findDataByKey(self,search_key):
        found_fields = set()
        for path in self.paths:
            print("Sprawdzam ",path, search_key)
            czyPlik = False
            try:
                if os.path.exists(self.paths_folder + path):
                    wb = openpyxl.load_workbook(self.paths_folder + path)
                    print("Plik otwarty poprawnie.")
                    czyPlik = True
                else:
                    print("Plik nie istnieje, idziemy dalej.")
            except Exception as e:
                    print(f"Nie udało się otworzyć pliku: {e}")
            
            if not czyPlik:
                print("zatrzymano")
                continue
            print("przeszlo")
            sheet = wb.active
            target_column = None

            print("Szukam klucza ", search_key)
            for index, cell in enumerate(sheet[1]):
                key = str(cell.value)
                print(key)
                if key == search_key:
                    target_column = index + 1
                    
                    print("Znaleziono klucz w pliku")
                    print("ta kolumna",target_column)
                elif key not in self.info.keys():
                    self.info[key] = None  # pozwalamy na dodatkowe pola
            
            print(self.info)
            print("sprawdzam kolumne")
            if target_column:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=target_column)
                    if str(cell.value) == self.info[search_key]:
                        print("znaleziono id")
                        target_row = row
                        for i in range(1, sheet.max_column + 1):
                            header = str(sheet.cell(row=1, column=i).value)
                            value = str(sheet.cell(row=target_row, column=i).value)
                            if not self.info.get(header):  # tylko jeśli nie ma wartości
                                self.info[header] = value
                                found_fields.add(header)
                            print("Znaleziono ", header, "z wartością", value)


                        break
        return found_fields


    def prepare_frame(self, parent, row, document_info, par=None):
        self.my_parent = par

        container = tk.Frame(parent, bg="#1e1e2f", highlightbackground="#666666", highlightcolor="white", highlightthickness=2)
        container.grid(row=row, column=0, sticky="wne", padx=10, pady=10)

        parent.grid_columnconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=0)  # info_container – nie rozszerza się
        container.grid_columnconfigure(1, weight=1)  # decison_container – nie rozszerza się
        container.grid_columnconfigure(2, weight=1)  # valuation_container – wypełnia resztę

        container.grid_rowconfigure(1, weight=1)
        container.grid_rowconfigure(2, weight=1)
        container.grid_rowconfigure(3, weight=1)

        device_label = ttk.Label(container, text=f"Urządzenie {row + 1}:")
        device_label.grid(row=0, column=0, sticky="nw", padx=10, pady=(10,2))

        def destroy_me():
            container.destroy()
            if self in self.my_parent.devices:
                self.my_parent.devices.remove(self)

        destroy_button = ttk.Button(container,
                                    text="X",
                                    style="Destroy2.TButton",
                                    command=destroy_me)
        destroy_button.grid(row=0, column=2, sticky="e", padx=10, pady=(10, 5))
        destroy_button.configure(width=2)

        info_container = tk.Frame(container, bg="#2e2e3e", highlightbackground="#666666", highlightcolor="white", highlightthickness=1)
        info_container.grid(row=1, column=0, sticky="wn", padx=10, pady=10, rowspan=3)

        decison_container = tk.Frame(container, bg="#2e2e3e", highlightbackground="#666666", highlightcolor="white", highlightthickness=1)
        decison_container.grid(row=1, column=1, sticky="nsew", padx=10, pady=10)

        state_container = tk.Frame(container, bg="#2e2e3e", highlightbackground="#666666", highlightcolor="white", highlightthickness=1)
        state_container.grid(row=2, column=1, sticky="nsew", padx=10, pady=10)

        change_container = tk.Frame(container, bg="#2e2e3e", highlightbackground="#666666", highlightcolor="white", highlightthickness=1)
        change_container.grid(row=3, column=1, sticky="nsew", padx=10, pady=10)

        valuation_container = tk.Frame(container, bg="#2e2e3e", highlightbackground="#666666", highlightcolor="white", highlightthickness=1)
        valuation_container.grid(row=1, column=2, sticky="nsew", padx=10, pady=10, rowspan=3)
        valuation_container.grid_columnconfigure(5, weight=1)  # valuation_container – wypełnia resztę


        self.fields = {}

        ### Funkcje
        def create_input(container, label_text, key, row_idx, col_idx=0, state="normal", colspan_entry=1,
                         only_float=False, only_int=False):
            label = tk.Label(container, text=label_text, bg="#2e2e3e", fg="white")
            label.grid(row=row_idx, column=col_idx, sticky="w", padx=10, pady=5)

            entry = ttk.Entry(container)

            if only_float:
                def validate_float_input(P):
                    if P in ("", "."):
                        return True
                    try:
                        float_val = float(P)
                        if "." in P:
                            decimals = P.split(".")[1]
                            return len(decimals) <= 2
                        return True
                    except ValueError:
                        return False

                vcmd = container.register(validate_float_input)
                entry.config(validate="key", validatecommand=(vcmd, "%P"))

            elif only_int:
                def validate_int_input(P):
                    if P == "":
                        return True
                    return P.isdigit()

                vcmd = container.register(validate_int_input)
                entry.config(validate="key", validatecommand=(vcmd, "%P"))

            entry.grid(row=row_idx, column=col_idx + 1, columnspan=colspan_entry, sticky="ew", padx=10, pady=5)
            

            if self.info.get(key):
                entry.insert(0, self.info[key])

            entry.config(state=state)
            self.fields[key] = entry
            setattr(self, f"field_{key}", entry)

        def create_text_widget(parent, state="normal", height=5, width=50):
            text_widget = tk.Text(
                parent,
                height=height,
                width=width,
                bg="#3a3a4a" if state == "normal" else "#292929",
                fg="white" if state == "normal" else "#666666",
                insertbackground="white",  # kursor
                selectbackground="#4a90e2",  # zaznaczenie
                font=("Segoe UI", 10),
                wrap="word",
                relief="flat",
                borderwidth=1
            )
            text_widget.configure(state=state)
            return text_widget

        def destroy_protocol():
            self.formula_area.delete("1.0", tk.END)
            self.formula_area.insert(tk.INSERT, self.destroy_formula)
            valuation_state("disabled")

        def valuation_destroy_protocol():
            self.formula_area.delete("1.0", tk.END)
            self.formula_area.insert(tk.INSERT, self.valuation_destroy_formula)
            valuation_state()

        def valuation_takeover_protocol():
            self.formula_area.delete("1.0", tk.END)
            self.formula_area.insert(tk.INSERT, self.valuation_takeover_formula)
            valuation_state()

        def state_good_protocol():
            self.state_area.delete("1.0", tk.END)
            self.state_area.insert(tk.INSERT, self.state_good_formula)

        def state_passable_protocol():
            self.state_area.delete("1.0", tk.END)
            self.state_area.insert(tk.INSERT, self.state_passable_formula)

        def change_cooling_protocol():
            self.change_area.delete("1.0", tk.END)
            self.change_area.insert(tk.INSERT, self.change_cooling_formula)

        def change_battery_protocol():
            self.change_area.delete("1.0", tk.END)
            self.change_area.insert(tk.INSERT, self.change_battery_formula)

        def change_cooling_and_batery_protocol():
            self.change_area.delete("1.0", tk.END)
            self.change_area.insert(tk.INSERT, self.change_cooling_and_battery_formula)

        def valuation_state(state="normal"):
            def update_widget_state(widget):
                try:
                    widget.configure(state=state)

                    # Jeśli to Text, zmień także kolory
                    if isinstance(widget, tk.Text):
                        if state == "normal":
                            widget.configure(bg="#3a3a4a", fg="white")
                        else:
                            widget.configure(bg="#292929", fg="#666666")

                except tk.TclError:
                    pass  # Nie każdy widget ma `state`

            for container in [valuation_container, state_container, change_container]:
                for widget in container.winfo_children():
                    update_widget_state(widget)

            if state == "disabled":
                toggle_cooling_fields(force_state="disabled")
                toggle_battery_fields(force_state="disabled")
            else:
                toggle_cooling_fields()
                toggle_battery_fields()

        def toggle_cooling_fields(force_state=None):
            # Jeśli force_state (np. 'disabled') jest przekazany, użyj go bez względu na checkbox
            if force_state:
                state = force_state
            else:
                state = "normal" if self.change_cooling.get() else "disabled"
            self.fields["Cena_cooling"].config(state=state)
            self.fields["Link_cooling"].config(state=state)

        def toggle_battery_fields(force_state=None):
            if force_state:
                state = force_state
            else:
                state = "normal" if self.change_batery.get() else "disabled"
            self.fields["Cena_battery"].config(state=state)
            self.fields["Link_battery"].config(state=state)

        def change_formula():
            if self.change_cooling.get() and self.change_batery.get():
                change_cooling_and_batery_protocol()
            elif self.change_cooling.get():
                change_cooling_protocol()
            elif self.change_batery.get():
                change_battery_protocol()
            else:
                self.change_area.delete("1.0", tk.END)
            toggle_cooling_fields()
            toggle_battery_fields()

        def format_name(full_name):
            parts = full_name.strip().split()

            if len(parts) >= 2:
                initial = parts[0][0].upper()
                surname = parts[-1].capitalize()
                return f"{initial}.{surname}"
            else:
                return full_name

        ### Informacje o sprzęcie
        create_input(info_container,"ID:", "Numer ewidencyjny", 0)
        create_input(info_container,"Nr inv:", "Nr inwentarzowy", 1)
        create_input(info_container,"Nr seryjny:", "Numer seryjny", 2)
        create_input(info_container,"Producent:", "Producent", 3)
        create_input(info_container,"Model:", "Model", 4)
        create_input(info_container,"Typ:", "Typ", 5)
        create_input(info_container,"Data zakupu:", "Data zakupu", 6)
        create_input(info_container,"SAP:", "Numer środka trwałego", 7)
        create_input(info_container,"Procesor:", "Model procesora", 8)
        label = tk.Label(info_container, text="Grafika", bg="#2e2e3e", fg="white")
        label.grid(row=9, column=0, sticky="w", padx=10, pady=5)
        entry = ttk.Entry(info_container)
        entry.insert(0, "Zintegrowana")
        entry.grid(row=9, column=1, sticky="ew", padx=10, pady=5)
        setattr(self, f"field_Grafika", entry)
        self.fields["Grafika"] = entry
        create_input(info_container,"Pamięć RAM:", "Pamięć RAM (GB)", 10)
        create_input(info_container,"Dysk twardy:", "Pamięć HDD (GB)", 11)
        create_input(info_container,"System operacyjny:", "System operacyjny", 12)


        ### Wybór decyzji
        decison_container.grid_columnconfigure(0, weight=0)
        decison_container.grid_columnconfigure(1, weight=0)
        decison_container.grid_columnconfigure(2, weight=1)
        decison_container.grid_rowconfigure(1, weight=1)
        self.formula_area = create_text_widget(decison_container, state="normal")
        self.formula_area.grid(row=1, column=0, columnspan=3, sticky='nsew', padx=10, pady=5)

        self.decision = tk.StringVar()
        r1 = Radiobutton(decison_container, text="Likwidacja", variable=self.decision, value="Likwidacja", command=destroy_protocol)
        r1.grid(row=0, column=0, sticky="nwe", padx=10, pady=5)
        r2 = Radiobutton(decison_container, text="Wycena (Likwidacja)", variable=self.decision, value="WycenaL", command=valuation_destroy_protocol)
        r2.grid(row=0, column=1, sticky="nwe", padx=10, pady=5)
        r3 = Radiobutton(decison_container, text="Wycena (Przekazanie)", variable=self.decision, value="WycenaP", command=valuation_takeover_protocol)
        r3.grid(row=0, column=2, sticky="nwe", padx=10, pady=5)
        self.decision.set("Likwidacja")

        ### Wycena sprzętu
        for i in range(10):
            create_input(valuation_container,"Cena:", f"Price{i+1}", i, 0, "disabled",colspan_entry=2, only_float=True)
            create_input(valuation_container, "Link:", f"Link{i+1}", i,3, "disabled",colspan_entry=2)

        label = tk.Label(state_container, text="Stan:", bg="#2e2e3e", fg="white")
        label.grid(row=0, column=0, sticky="w", padx=10, pady=5)
        self.device_state = tk.StringVar()
        r1 = Radiobutton(state_container, text="Dobry", variable=self.device_state, value="Good", state="disabled", command=state_good_protocol)
        r1.grid(row=0, column=1, sticky="w", padx=10, pady=5)
        r2 = Radiobutton(state_container, text="Dostateczny", variable=self.device_state, value="Passable", state="disabled", command=state_passable_protocol)
        r2.grid(row=0, column=2, sticky="w", padx=10, pady=5)
        self.device_state.set("Good")
        setattr(self, f"radio_state_good", r1)
        setattr(self, f"radio_state_passable", r2)
        state_container.grid_columnconfigure(0, weight=0)
        state_container.grid_columnconfigure(1, weight=0)
        state_container.grid_columnconfigure(2, weight=1)
        state_container.grid_rowconfigure(1, weight=1)
        self.state_area = create_text_widget(state_container, state="normal", height=3)
        self.state_area.grid(row=1, column=0, columnspan=3, sticky='nsew', padx=10, pady=5)
        state_good_protocol()
        self.state_area.config(state="disabled")

        label = tk.Label(change_container, text="Wymiana:", bg="#2e2e3e", fg="white")
        label.grid(row=0, column=0, sticky="w", padx=10, pady=5)
        self.change_cooling = tk.BooleanVar(value=True)
        self.change_batery = tk.BooleanVar(value=True)
        c1 = Checkbutton(change_container, text="Chłodzenie", variable=self.change_cooling, onvalue=True, offvalue=False, state="disabled", command=change_formula)
        c1.grid(row=1, column=0, sticky="w", padx=10)
        create_input(change_container, "Cena:", "Cena_cooling", 1, 1, state="disabled", only_float=True)
        create_input(change_container, "Link:", "Link_cooling", 1, 3, state="disabled")
        c2 = Checkbutton(change_container, text="Bateria", variable=self.change_batery, onvalue=True, offvalue=False, state="disabled", command=change_formula)
        create_input(change_container, "Cena:", "Cena_battery", 2, 1, state="disabled", only_float=True)
        create_input(change_container, "Link:", "Link_battery", 2, 3, state="disabled")
        c2.grid(row=2, column=0, sticky="w", padx=10)
        setattr(self, f"checkbox_cooling", c1)
        setattr(self, f"checkbox_battery", c2)
        change_container.grid_columnconfigure(0, weight=0)
        change_container.grid_columnconfigure(1, weight=0)
        change_container.grid_columnconfigure(2, weight=0)
        change_container.grid_columnconfigure(3, weight=0)
        change_container.grid_columnconfigure(4, weight=1)
        change_container.grid_rowconfigure(3, weight=1)
        self.change_area = create_text_widget(change_container, state="normal", height=3)
        self.change_area.grid(row=3, column=0, columnspan=5, sticky='nsew', padx=10, pady=5)
        change_cooling_and_batery_protocol()
        self.change_area.config(state="disabled")

        create_input(valuation_container, "MPK:", "MPK", 12, colspan_entry=5, state="disabled")
        create_input(valuation_container, "Nr wyceny:", "Numer wyceny", 13, colspan_entry=2, state="disabled", only_int=True)

        label = tk.Label(valuation_container, text="Nazwa:", bg="#2e2e3e", fg="white")
        label.grid(row=14, column=0, sticky="w", padx=10, pady=5)
        reporter_name = format_name(document_info["Zglaszajacy"])
        self.pricing_file_name = f"Formularz - Wycena sprzętu komputerowego ID_{self.info['Numer ewidencyjny']}_{reporter_name}"
        entry = ttk.Entry(valuation_container)
        entry.insert(0, self.pricing_file_name)
        entry.grid(row=14, column=1, columnspan=5, sticky="ew", padx=10, pady=5)
        entry.config(state="disabled")
        setattr(self, f"field_pricing_file_name", entry)
        self.fields["pricing_file_name"] = entry


        ### Protokoły
        destroy_protocol()
        self.updateDeviceInfo()

        return container, fields

    def getDecision(self):
        if self.decision.get() == "Likwidacja":
            return False
        return True

    def validateRequiredFields(self):
        required_fields = [
            "Numer ewidencyjny", "Nr inwentarzowy", "Numer seryjny", "Producent",
            "Model", "Typ", "Data zakupu", "Numer środka trwałego", "Model procesora",
            "Pamięć RAM (GB)", "Pamięć HDD (GB)", "System operacyjny", "Numer wyceny"
        ]

        missing = []
        for field in required_fields:
            widget = self.fields.get(field)
            if not widget:
                continue  # brak widgetu – pomiń

            # NIE sprawdzamy wartości, jeśli pole jest disabled
            try:
                if str(widget["state"]) == "disabled":
                    continue
            except:
                pass  # jeśli widget nie ma 'state', ignoruj

            # Jeśli pole aktywne i puste → dodaj do brakujących
            if not widget.get().strip():
                missing.append(field)

        if missing:
            msg = "Brakuje danych w następujących polach:\n\n" + "\n".join(missing)
            msg += "\n\nCzy chcesz kontynuować mimo braków?"
            return messagebox.askyesno("Niekompletne dane urządzenia", msg)

        return True

    def updateDeviceInfo(self):
        for field, value in self.fields.items():
            if value.cget("state") != "disabled":
                self.info[field] = value.get()

        self.info["Formula"] = self.formula_area.get(1.0, "end-1c")

        if self.state_area.cget("state") != "disabled":
            self.info["Stan"] = self.state_area.get(1.0, "end-1c")
        else:
            self.info["Stan"] = ""

        if self.change_area.cget("state") != "disabled":
            self.info["Wymiana"] = self.change_area.get(1.0, "end-1c")
        else:
            self.info["Wymiana"] = ""

        if not self.change_cooling.get():
            self.info["Cena_cooling"] = ""
            self.info["Link_cooling"] = ""

        if not self.change_batery.get():
            self.info["Cena_battery"] = ""
            self.info["Link_battery"] = ""